from io import BytesIO
from pathlib import Path

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from tortoise.expressions import Q
from tortoise.transactions import in_transaction

from app.core.crud import CRUDBase
from app.exam.constants import EXAM_IMPORT_HEADERS, EXAM_IMPORT_SHEET_NAME
from app.exam.models import ExamBank, ExamPaperQuestion, ExamQuestion
from app.exam.schemas.bank import ExamBankCreate, ExamBankUpdate
from app.exam.storage import remove_stored_file, save_bank_source_file
from app.exam.utils import (
    build_import_row_payload,
    normalize_question_payload,
    normalize_text,
    validate_import_headers,
)


class ExamBankController(CRUDBase[ExamBank, ExamBankCreate, ExamBankUpdate]):
    def __init__(self):
        super().__init__(model=ExamBank)

    async def list_banks(self, page: int, page_size: int, name: str | None = None, is_active: bool | None = None):
        q = Q()
        if name:
            q &= Q(name__contains=name)
        if is_active is not None:
            q &= Q(is_active=is_active)

        total, bank_objs = await self.list(page=page, page_size=page_size, search=q, order=["-id"])
        data = [await obj.to_dict() for obj in bank_objs]
        return total, data

    async def get_bank_detail(self, bank_id: int):
        bank = await self.get(id=bank_id)
        data = await bank.to_dict()
        data["question_count"] = await ExamQuestion.filter(bank_id=bank.id).count()
        return data

    async def create_bank(self, obj_in: ExamBankCreate, user_id: int):
        exists = await self.model.filter(name=obj_in.name).exists()
        if exists:
            raise HTTPException(status_code=400, detail="题库名称已存在")

        payload = obj_in.model_dump()
        payload["created_by"] = user_id
        payload["updated_by"] = user_id
        return await self.create(obj_in=payload)

    async def update_bank(self, obj_in: ExamBankUpdate, user_id: int):
        await self.get(id=obj_in.id)
        exists = await self.model.filter(name=obj_in.name).exclude(id=obj_in.id).exists()
        if exists:
            raise HTTPException(status_code=400, detail="题库名称已存在")

        payload = obj_in.model_dump()
        payload["updated_by"] = user_id
        return await self.update(id=obj_in.id, obj_in=payload)

    async def delete_bank(self, bank_id: int):
        bank = await self.get(id=bank_id)
        question_ids = await ExamQuestion.filter(bank_id=bank_id).values_list("id", flat=True)
        if question_ids:
            referenced = await ExamPaperQuestion.filter(question_id__in=list(question_ids)).exists()
            if referenced:
                raise HTTPException(status_code=400, detail="题库中的题目已被试卷引用，无法删除")
            await ExamQuestion.filter(bank_id=bank_id).delete()

        remove_stored_file(bank.source_file_path)
        await bank.delete()

    def build_template_file(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = EXAM_IMPORT_SHEET_NAME
        worksheet.append(EXAM_IMPORT_HEADERS)
        worksheet.append(
            [
                "单选题",
                "Python/基础",
                "Python 中列表使用哪个符号定义？",
                "[]",
                "{}",
                "()",
                "<>",
                "",
                "",
                "A",
                5,
                "简单",
                "列表使用方括号定义",
                "",
                "Python,列表",
            ]
        )
        worksheet.append(
            [
                "多选题",
                "Python/基础",
                "以下哪些是 Python 内置数据类型？",
                "list",
                "dict",
                "array",
                "tuple",
                "",
                "",
                "A,B,D",
                5,
                "中等",
                "list、dict、tuple 都是 Python 内置数据类型",
                "",
                "Python,数据类型",
            ]
        )
        worksheet.append(
            [
                "判断题",
                "FastAPI/基础",
                "FastAPI 属于 Python Web 框架。",
                "",
                "",
                "",
                "",
                "",
                "",
                "正确",
                5,
                "简单",
                "FastAPI 是基于 Python 的 Web 框架",
                "",
                "FastAPI",
            ]
        )
        worksheet.append(
            [
                "填空题",
                "SQL/基础",
                "请写出查询全部字段的 SQL 语句。",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                10,
                "简单",
                "",
                "SELECT * FROM table_name;",
                "SQL",
            ]
        )
        worksheet.append(
            [
                "简答题",
                "系统设计/基础",
                "请简述缓存穿透的含义及常见解决方案。",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                10,
                "困难",
                "",
                "缓存穿透指查询不存在的数据导致请求直接打到数据库，常见方案包括布隆过滤器、缓存空值等。",
                "系统设计,缓存",
            ]
        )
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    async def import_bank_questions(
        self,
        name: str,
        desc: str | None,
        is_active: bool,
        file: UploadFile,
        user_id: int,
    ):
        if await self.model.filter(name=name).exists():
            return {
                "success": False,
                "errors": [{"row": 0, "field": "题库名称", "message": "题库名称已存在"}],
            }

        file_name = file.filename or ""
        if Path(file_name).suffix.lower() != ".xlsx":
            return {
                "success": False,
                "errors": [{"row": 0, "field": "文件", "message": "仅支持上传 .xlsx 文件"}],
            }

        file_bytes = await file.read()
        if not file_bytes:
            return {
                "success": False,
                "errors": [{"row": 0, "field": "文件", "message": "上传文件不能为空"}],
            }

        try:
            workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        except Exception:
            return {
                "success": False,
                "errors": [{"row": 0, "field": "文件", "message": "题库文件解析失败，请检查文件格式"}],
            }

        if EXAM_IMPORT_SHEET_NAME not in workbook.sheetnames:
            return {
                "success": False,
                "errors": [
                    {
                        "row": 0,
                        "field": "Sheet",
                        "message": f"题库模板必须包含名为 {EXAM_IMPORT_SHEET_NAME} 的 Sheet",
                    }
                ],
            }

        worksheet = workbook[EXAM_IMPORT_SHEET_NAME]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [normalize_text(value) for value in (header_row or [])]

        try:
            validate_import_headers(headers)
        except HTTPException as exc:
            return {
                "success": False,
                "errors": [{"row": 1, "field": "表头", "message": str(exc.detail)}],
            }

        header_index_map = {header: index for index, header in enumerate(headers)}
        seen_stems = set()
        errors = []
        question_payloads = []

        for row_number, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {}
            for header in EXAM_IMPORT_HEADERS:
                index = header_index_map.get(header)
                row_data[header] = row_values[index] if index is not None and index < len(row_values) else None

            if all(not normalize_text(value) for value in row_data.values()):
                continue

            stem = normalize_text(row_data.get("题干"))
            if stem and stem in seen_stems:
                errors.append({"row": row_number, "field": "题干", "message": "同一题库中题干重复"})
                continue
            if stem:
                seen_stems.add(stem)

            try:
                payload = build_import_row_payload(row_data)
                payload["bank_id"] = 0
                normalized = normalize_question_payload(payload)
                question_payloads.append(normalized)
            except HTTPException as exc:
                errors.append({"row": row_number, "field": "题目", "message": str(exc.detail)})

        if not question_payloads and not errors:
            errors.append({"row": 0, "field": "题库", "message": "题库中没有可导入的有效题目"})

        if errors:
            return {"success": False, "errors": errors}

        stored_file = save_bank_source_file(file_name=file_name, file_bytes=file_bytes)

        async with in_transaction() as connection:
            bank = await ExamBank.create(
                name=name,
                desc=desc,
                source_file_name=stored_file["file_name"],
                source_file_path=stored_file["file_path"],
                question_count=len(question_payloads),
                is_active=is_active,
                created_by=user_id,
                updated_by=user_id,
                using_db=connection,
            )
            for payload in question_payloads:
                payload["bank_id"] = bank.id
                payload["created_by"] = user_id
                payload["updated_by"] = user_id
                await ExamQuestion.create(**payload, using_db=connection)

        return {
            "success": True,
            "bank_id": bank.id,
            "question_count": len(question_payloads),
        }


exam_bank_controller = ExamBankController()
