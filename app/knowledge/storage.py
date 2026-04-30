import json
import mimetypes
import shutil
import uuid
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

import mammoth
from fastapi import HTTPException, UploadFile
from markdown_it import MarkdownIt
from openpyxl import load_workbook

from app.knowledge.constants import (
    KNOWLEDGE_ALLOWED_FILE_EXTENSIONS,
    KnowledgeArticleBlockType,
)
from app.settings import settings

MARKDOWN_RENDERER = MarkdownIt("commonmark", {"html": False, "linkify": True})
KNOWLEDGE_STORAGE_DIR = Path(settings.BASE_DIR) / "storage" / "knowledge"
KNOWLEDGE_TEMP_DIR = KNOWLEDGE_STORAGE_DIR / "_tmp"
KNOWLEDGE_ARTICLE_DIR = KNOWLEDGE_STORAGE_DIR / "articles"
MAX_UPLOAD_SIZE = 30 * 1024 * 1024
MAX_EXCEL_PREVIEW_ROWS = 200
TEMP_UPLOAD_TTL = timedelta(hours=24)


def _ensure_storage_dirs():
    KNOWLEDGE_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    KNOWLEDGE_ARTICLE_DIR.mkdir(parents=True, exist_ok=True)


def _to_relative_storage_path(path: Path) -> str:
    return path.relative_to(settings.BASE_DIR).as_posix()


def get_storage_abspath(relative_path: str) -> Path:
    return Path(settings.BASE_DIR) / relative_path


def cleanup_temp_uploads():
    _ensure_storage_dirs()
    expire_before = datetime.now() - TEMP_UPLOAD_TTL
    for path in KNOWLEDGE_TEMP_DIR.iterdir():
        if not path.is_file():
            continue
        modified_at = datetime.fromtimestamp(path.stat().st_mtime)
        if modified_at >= expire_before:
            continue
        path.unlink(missing_ok=True)


def _normalize_cell_value(value: Any):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime(settings.DATETIME_FORMAT)
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, Decimal):
        return float(value)
    return value


def _trim_row(row: list[Any]):
    while row and row[-1] == "":
        row.pop()
    return row


def _detect_block_type(filename: str):
    ext = Path(filename).suffix.lower()
    if ext not in KNOWLEDGE_ALLOWED_FILE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持上传 .md / .docx / .pdf / .xlsx 文件")
    return KnowledgeArticleBlockType(ext.lstrip(".")), ext


def _decode_markdown_bytes(file_bytes: bytes):
    encodings = ("utf-8-sig", "utf-8", "gbk")
    for encoding in encodings:
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=400, detail="Markdown 文件编码无法识别，请使用 UTF-8 或 GBK")


def _build_excel_preview(file_bytes: bytes):
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    sheets = []

    try:
        for worksheet in workbook.worksheets:
            rows = []
            max_column_count = 0
            truncated = False

            for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                normalized_row = _trim_row([_normalize_cell_value(value) for value in row])
                if normalized_row:
                    rows.append(normalized_row)
                    max_column_count = max(max_column_count, len(normalized_row))
                if index >= MAX_EXCEL_PREVIEW_ROWS:
                    truncated = True
                    break

            sheets.append(
                {
                    "name": worksheet.title,
                    "rows": rows,
                    "max_column_count": max_column_count,
                    "truncated": truncated,
                }
            )
    finally:
        workbook.close()

    return {
        "sheets": sheets,
        "sheet_count": len(sheets),
        "preview_row_limit": MAX_EXCEL_PREVIEW_ROWS,
    }


def _build_render_payload(block_type: KnowledgeArticleBlockType, file_bytes: bytes):
    render_html = None
    render_json = None
    extra_meta = {}

    if block_type == KnowledgeArticleBlockType.MARKDOWN:
        markdown_text = _decode_markdown_bytes(file_bytes)
        render_html = MARKDOWN_RENDERER.render(markdown_text)
        extra_meta["line_count"] = markdown_text.count("\n") + 1 if markdown_text else 0
    elif block_type == KnowledgeArticleBlockType.DOCX:
        result = mammoth.convert_to_html(BytesIO(file_bytes))
        render_html = result.value
        if result.messages:
            extra_meta["messages"] = [message.message for message in result.messages]
    elif block_type == KnowledgeArticleBlockType.XLSX:
        render_json = _build_excel_preview(file_bytes)

    return render_html, render_json, extra_meta or None


async def save_temp_upload(file: UploadFile):
    cleanup_temp_uploads()
    _ensure_storage_dirs()

    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    block_type, ext = _detect_block_type(file.filename)
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="上传文件不能为空")
    if len(file_bytes) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="单个文件大小不能超过 30MB")

    render_html, render_json, extra_meta = _build_render_payload(block_type, file_bytes)
    temp_token = uuid.uuid4().hex
    temp_file_path = KNOWLEDGE_TEMP_DIR / f"{temp_token}{ext}"
    temp_meta_path = KNOWLEDGE_TEMP_DIR / f"{temp_token}.json"

    temp_file_path.write_bytes(file_bytes)

    payload = {
        "temp_token": temp_token,
        "block_type": block_type.value,
        "file_name": file.filename,
        "file_ext": ext,
        "mime_type": file.content_type or mimetypes.guess_type(file.filename)[0] or "application/octet-stream",
        "file_size": len(file_bytes),
        "file_path": _to_relative_storage_path(temp_file_path),
        "render_html": render_html,
        "render_json": render_json,
        "extra_meta": extra_meta,
    }
    temp_meta_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return payload


def load_temp_upload(temp_token: str):
    cleanup_temp_uploads()
    temp_meta_path = KNOWLEDGE_TEMP_DIR / f"{temp_token}.json"
    if not temp_meta_path.exists():
        raise HTTPException(status_code=404, detail="上传文件已失效，请重新上传")

    data = json.loads(temp_meta_path.read_text(encoding="utf-8"))
    file_path = get_storage_abspath(data["file_path"])
    if not file_path.exists():
        temp_meta_path.unlink(missing_ok=True)
        raise HTTPException(status_code=404, detail="上传文件不存在，请重新上传")
    return data


def finalize_temp_upload(temp_token: str, article_id: int):
    temp_data = load_temp_upload(temp_token)
    source_path = get_storage_abspath(temp_data["file_path"])
    target_dir = KNOWLEDGE_ARTICLE_DIR / str(article_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / f"{uuid.uuid4().hex}{temp_data['file_ext']}"
    shutil.move(str(source_path), str(target_path))

    temp_meta_path = KNOWLEDGE_TEMP_DIR / f"{temp_token}.json"
    temp_meta_path.unlink(missing_ok=True)

    return {
        "block_type": temp_data["block_type"],
        "file_name": temp_data["file_name"],
        "file_ext": temp_data["file_ext"],
        "mime_type": temp_data["mime_type"],
        "file_size": temp_data["file_size"],
        "file_path": _to_relative_storage_path(target_path),
        "render_html": temp_data.get("render_html"),
        "render_json": temp_data.get("render_json"),
        "extra_meta": temp_data.get("extra_meta"),
    }


def remove_stored_file(relative_path: str | None):
    if not relative_path:
        return

    file_path = get_storage_abspath(relative_path)
    file_path.unlink(missing_ok=True)

    parent = file_path.parent
    while parent != KNOWLEDGE_STORAGE_DIR and parent.exists():
        try:
            parent.rmdir()
        except OSError:
            break
        parent = parent.parent
